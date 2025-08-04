from openpyxl import load_workbook
from openpyxl import Workbook

def file_dir (year, month):
    download_dir = f"C:\\dev\\mini_1\\data\\{year}\\" 
    return download_dir + f'{year}년_{month}월_자동차_등록자료_통계.xlsx'

def copy_data(ws, ad_ws):
    # 안전한 데이터 복사
    for row_num in range(1, ws.max_row + 1):
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value is not None:  # 빈 셀이 아닌 경우만 복사
                ad_ws.cell(row=row_num, column=col_num, value=cell.value)

def remove_row(ws_edit):    
    # 남겨둘 행들 (1부터 시작하는 인덱스)
    rows_to_keep = [3,13,23]

    # 제거할 행들 찾기
    rows_to_remove = []
    for row in range(ws_edit.max_row, 0, -1):  # 뒤에서부터 확인
        if row not in rows_to_keep:
            rows_to_remove.append(row)
    # 행행 제거
    for row in rows_to_remove:
        if row <= ws_edit.max_row:
            ws_edit.delete_rows(row)


def remove_col(ws_edit):
    #열 제거
    columns_to_remove = [3,2]
    for column in columns_to_remove:
        ws_edit.delete_cols(column)    

sheet_number = 3
row_number = 1
for year in range(2024,2016, -1):
    for month in range(1,13):
        dir =file_dir(year, month) 
        wb = load_workbook(dir, read_only=True)  # 복사할 파일 열기
        ws=wb[wb.sheetnames[sheet_number-1]] #1번시트
        new_wb = Workbook() # 저장할 새 파일 생성
        ad_ws = new_wb.active
        ad_ws.title = wb.sheetnames[sheet_number-1]   
        copy_data(ws, ad_ws)
        new_file_name = dir.replace('.xlsx', f'_{sheet_number}.xlsx') # 새 파일로 저장
        new_wb.save(new_file_name)
        wb_edit = load_workbook(new_file_name) # 편집 가능한 파일로 열기
        ws_edit = wb_edit.active
        remove_col(ws_edit)
        remove_row(ws_edit)
        ws_edit.cell(row=2, column=1, value='남자')
        ws_edit.cell(row=3, column=1, value='여자')
        wb_edit.save(new_file_name) # 변경사항을 반드시 저장
        print(f"{year},{month}변경사항 저장 완료")

