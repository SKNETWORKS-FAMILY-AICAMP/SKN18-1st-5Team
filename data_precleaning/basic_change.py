import os
from openpyxl import load_workbook

# 파일 이름 바꾸기
def rename_file():
    for year in range(2021, 2025):
        base_directory = f"C:\\dev\\mini_1\\data\\{year}\\"
        for month in range(1, 10):
            old_name = base_directory + f'{year}년_0{month}월_자동차_등록자료_통계.xlsx'
            new_name = base_directory + f'{year}년_{month}월_자동차_등록자료_통계.xlsx'
        
            if os.path.exists(old_name):
                os.rename(old_name, new_name)
                print(f"파일명 변경: {old_name} → {new_name}")
            else:
                print(f"파일 없음: {old_name}")
                
# 시트 버리기                
def drop_sheet():
    # 남길 시트 이름
    keep_sheets = ['01.통계표', '02.통계표_시군구' ,'04.성별_연령별']     
    ############ 연도 기간###########
    START_YEAR = 2021
    END_YEAR = 2021
    ############ 달 기간 ################
    START_MONTH = 8
    END_MONTH = 8
    for year in range(START_YEAR, END_YEAR + 1):
        base_directory = f"C:\\dev\\mini_1\\data\\{year}\\"
        for month in range(START_MONTH, END_MONTH+1):
            # 원본 엑셀 파일 경로        
            file_path = base_directory+ f'{year}년_{month}월_자동차_등록자료_통계.xlsx'
            if not os.path.exists(file_path):
                print(f"파일 없음: {file_path}")
                continue  # 파일이 없으면 아래 코드를 건너뜀
            wb = load_workbook(file_path) # 엑셀 파일 열기
            for sheet in wb.sheetnames: # 남길 시트를 제외한 모든 시트 삭제
                if sheet not in keep_sheets:
                    std = wb[sheet]
                    wb.remove(std)
            wb.save(file_path) # 덮어쓰기 저장 (백업 권장)
            print(f"덮어쓰기 완료 {year}{month}")